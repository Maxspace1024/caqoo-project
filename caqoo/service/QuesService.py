from bean.QuesListBean import *

from sqlite3 import Connection, Cursor
from fastapi import UploadFile

import openpyxl

class QuesService():
    conn : Connection
    cur : Cursor

    def __init__(self, conn : Connection):
        self.conn = conn
        self.cur = conn.cursor()

    async def getQuesesById(self, id : int) -> QuesListBean: 
        res = self.cur.execute(f"select * from question where id_question_file = {id}")
        resBean = []
        for obj in res.fetchall():
            obj = dict(obj)
            resBean.append(QuesBean(
                obj['id'],
                obj['ques'], 
                [obj["ans0"], obj["ans1"], obj["ans2"], obj["ans3"]], 
                obj["true_ans"], 
                obj["ans_detail"]
                )
            )

        return QuesListBean(id, resBean)
    
    async def saveQueses(self, file : UploadFile):
        # 新增question_file
        self.cur.execute(f"insert into question_file (filename) values ('{file.filename}')")
        self.conn.commit()
        newrowid = self.cur.lastrowid

        # 讀取xlsx file
        wb = openpyxl.open(filename=file.file)
        ws = wb["工作表1"]
        rowIndex = 2
        ans0 = ws.cell(row=rowIndex, column=1).value
        while ans0 != None and rowIndex<15:
            ans1 = ws.cell(row=rowIndex, column=2).value
            ans2 = ws.cell(row=rowIndex, column=3).value
            ans3 = ws.cell(row=rowIndex, column=4).value
            trueAns = ws.cell(row=rowIndex, column=5).value
            ansDetail = ws.cell(row=rowIndex, column=6).value
            ques = ws.cell(row=rowIndex, column=7).value
            # print(ans0,ans1,ans2,ans3, trueAns, ansDetail, ques)
            # 將題目寫入question
            self.cur.execute(f"insert into question (ans0,ans1,ans2,ans3, true_ans, ans_detail, ques, id_question_file) values ('{ans0}','{ans1}','{ans2}','{ans3}', {trueAns}, '{ansDetail}', '{ques}', {newrowid})")
            self.conn.commit()

            rowIndex+=1
            ans0 = ws.cell(row=rowIndex, column=1).value
        wb.close()

    async def getQuesesAll(self):
        res = self.cur.execute(f"select * from question_file")
        resBean = []
        for obj in res.fetchall():
            resBean.append(dict(obj))
        return resBean
    
    async def getQueses(self, id : int): 
        res = self.cur.execute(f"select * from question where id_question_file = {id}")
        resBean = []
        for obj in res.fetchall():
            resBean.append(dict(obj))
        return resBean
